# External packages
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference

def new_file():
    # Load file
    # Set file path (please give name of the file for example: Exampledata)
    # This syntax will work if your existing excel file is in your current working directory. 
    # In case if you want to access a spreadsheet from anywhere on your computer you will provide a path to your file with full file name and extension. 
    workbook_name = input('Enter the name of the workbook: ')
    wb2 = load_workbook(filename=workbook_name + ".xlsx")
    ## alternative way of loading file:
    ## filepath = "/Users/katarzynaiwaszkiewicz/Desktop/projekt_python/projekt/report_pivot.xlsx"
    ## wb2 = load_workbook(filepath)
    sheet = wb2.active
    # Define labels from no. of columns and rows
    labels = Reference(sheet, min_col=1, min_row=2, max_row=8)
    # Define numeric values from no. of columns and rows
    values = Reference(sheet, min_col=2, max_col=5, min_row=1, max_row=8)

    chart = BarChart()
    # Define height and width of a chart
    chart.height = 20
    chart.width = 30
    chart.add_data(values, titles_from_data=True)
    # Define a title of a chart
    chart.title="Przychody firm"
    chart.set_categories(labels)
    # Define titles of x and y axis
    chart.x_axis.title = "Firmy"
    chart.y_axis.title = "Przychód"
    # Choose cell in which the chart will be displayed.
    sheet.add_chart(chart,"A13")
    # Choose file name
    wb2.save("BarChart.xlsx")

new_file()